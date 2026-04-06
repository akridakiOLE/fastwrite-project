"""
Module 7: Μηχανισμός Εξαγωγής (Export & Search)
Λειτουργίες:
  1. Αναζήτηση / φιλτράρισμα επεξεργασμένων εγγράφων
  2. Εξαγωγή σε CSV  (τοπικά)
  3. Εξαγωγή σε XLSX (τοπικά, με formatting)
  4. Εξαγωγή μεμονωμένου εγγράφου σε JSON
"""

import csv
import json
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd


# ── Σταθερές ─────────────────────────────────────────────────────────────────

DEFAULT_EXPORT_DIR = Path("/app/projects/exports")

# Στήλες που εμφανίζονται πάντα πρώτες στην εξαγωγή
PRIORITY_COLUMNS = [
    "id", "filename", "status", "invoice_number", "invoice_date",
    "vendor_afm", "buyer_afm", "net_amount", "vat_rate",
    "vat_amount", "total_amount", "discount", "created_at",
]


# ── Dataclasses ───────────────────────────────────────────────────────────────

from dataclasses import dataclass, field as dc_field

@dataclass
class ExportResult:
    """Αποτέλεσμα εξαγωγής."""
    success      : bool       = False
    file_path    : Path       = None
    format       : str        = ""        # "csv" | "xlsx" | "json"
    record_count : int        = 0
    error        : str        = ""
    exported_at  : str        = ""

    def to_dict(self) -> Dict:
        return {
            "success":      self.success,
            "file_path":    str(self.file_path) if self.file_path else None,
            "format":       self.format,
            "record_count": self.record_count,
            "error":        self.error,
            "exported_at":  self.exported_at,
        }


@dataclass
class SearchResult:
    """Αποτέλεσμα αναζήτησης."""
    records      : List[Dict[str, Any]] = dc_field(default_factory=list)
    total_count  : int                  = 0
    query        : str                  = ""
    filters_used : Dict                 = dc_field(default_factory=dict)


# ── Κύρια Κλάση ───────────────────────────────────────────────────────────────

class DocumentExporter:
    """
    Εξαγωγή και αναζήτηση επεξεργασμένων εγγράφων.

    Δέχεται λίστα από dicts (records) που προέρχονται από:
      - DatabaseManager.list_documents()   (metadata)
      - AIExtractor.extracted_data         (περιεχόμενο)
    """

    def __init__(self, export_dir: Path = None):
        self.export_dir = Path(export_dir) if export_dir else DEFAULT_EXPORT_DIR
        self.export_dir.mkdir(parents=True, exist_ok=True)

    # ── Αναζήτηση ────────────────────────────────────────────────────────────

    def search(self, records: List[Dict[str, Any]],
               query: str = "",
               status_filter: str = None,
               date_from: str = None,
               date_to: str = None,
               min_amount: float = None,
               max_amount: float = None) -> SearchResult:
        """
        Φιλτράρει records βάσει κριτηρίων.

        :param records:       Λίστα εγγράφων (dicts)
        :param query:         Ελεύθερο κείμενο (ψάχνει σε filename, invoice_number, afm)
        :param status_filter: π.χ. "Validated" | "Needs Human Review" | "Pending"
        :param date_from:     ISO date "YYYY-MM-DD" — από
        :param date_to:       ISO date "YYYY-MM-DD" — έως
        :param min_amount:    Ελάχιστο total_amount
        :param max_amount:    Μέγιστο total_amount
        """
        result = SearchResult(
            query=query,
            filters_used={k: v for k, v in {
                "status": status_filter, "date_from": date_from,
                "date_to": date_to, "min_amount": min_amount,
                "max_amount": max_amount,
            }.items() if v is not None}
        )

        filtered = list(records)

        # ── Φίλτρο κειμένου ─────────────────────────────────────────────
        if query and query.strip():
            q = query.strip().lower()
            SEARCH_FIELDS = ["filename", "invoice_number",
                             "vendor_afm", "buyer_afm", "status"]
            filtered = [
                r for r in filtered
                if any(q in str(r.get(f, "")).lower() for f in SEARCH_FIELDS)
            ]

        # ── Φίλτρο status ────────────────────────────────────────────────
        if status_filter:
            filtered = [
                r for r in filtered
                if r.get("status", "").lower() == status_filter.lower()
            ]

        # ── Φίλτρο ημερομηνίας ──────────────────────────────────────────
        if date_from:
            filtered = [
                r for r in filtered
                if r.get("invoice_date", "") >= date_from
            ]
        if date_to:
            filtered = [
                r for r in filtered
                if r.get("invoice_date", "") <= date_to
            ]

        # ── Φίλτρο ποσού ────────────────────────────────────────────────
        if min_amount is not None:
            filtered = [
                r for r in filtered
                if isinstance(r.get("total_amount"), (int, float))
                and r["total_amount"] >= min_amount
            ]
        if max_amount is not None:
            filtered = [
                r for r in filtered
                if isinstance(r.get("total_amount"), (int, float))
                and r["total_amount"] <= max_amount
            ]

        result.records     = filtered
        result.total_count = len(filtered)
        return result

    # ── Εξαγωγή CSV ──────────────────────────────────────────────────────────

    def export_csv(self, records: List[Dict[str, Any]],
                   filename: str = None,
                   columns: List[str] = None) -> ExportResult:
        """
        Εξαγωγή records σε CSV αρχείο.

        :param records:  Λίστα dicts
        :param filename: Όνομα αρχείου (auto-generated αν None)
        :param columns:  Λίστα στηλών (όλες αν None)
        """
        result = ExportResult(format="csv")

        if not records:
            result.error = "Δεν υπάρχουν δεδομένα για εξαγωγή."
            return result

        try:
            df       = self._to_dataframe(records, columns)
            out_path = self._make_output_path(filename, "csv")

            df.to_csv(str(out_path), index=False, encoding="utf-8-sig")
            # utf-8-sig: BOM για σωστή εμφάνιση ελληνικών στο Excel

            result.success      = True
            result.file_path    = out_path
            result.record_count = len(df)
            result.exported_at  = datetime.utcnow().isoformat()

        except Exception as e:
            result.error = str(e)

        return result

    # ── Εξαγωγή XLSX ─────────────────────────────────────────────────────────

    def export_xlsx(self, records: List[Dict[str, Any]],
                    filename: str = None,
                    columns: List[str] = None,
                    sheet_name: str = "Έγγραφα") -> ExportResult:
        """
        Εξαγωγή records σε Excel (.xlsx) με βασικό formatting.

        :param records:    Λίστα dicts
        :param filename:   Όνομα αρχείου (auto-generated αν None)
        :param columns:    Λίστα στηλών (όλες αν None)
        :param sheet_name: Όνομα φύλλου Excel
        """
        result = ExportResult(format="xlsx")

        if not records:
            result.error = "Δεν υπάρχουν δεδομένα για εξαγωγή."
            return result

        try:
            df       = self._to_dataframe(records, columns)
            out_path = self._make_output_path(filename, "xlsx")

            with pd.ExcelWriter(str(out_path), engine="openpyxl") as writer:
                df.to_excel(writer, index=False, sheet_name=sheet_name)

                # ── Formatting ───────────────────────────────────────────
                wb = writer.book
                ws = writer.sheets[sheet_name]

                # Header styling
                from openpyxl.styles import Font, PatternFill, Alignment
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(
                    start_color="2563EB", end_color="2563EB",
                    fill_type="solid"
                )
                for cell in ws[1]:
                    cell.font      = header_font
                    cell.fill      = header_fill
                    cell.alignment = Alignment(horizontal="center")

                # Auto-width στηλών
                for col in ws.columns:
                    max_len = max(
                        (len(str(cell.value)) if cell.value else 0)
                        for cell in col
                    )
                    ws.column_dimensions[col[0].column_letter].width = (
                        min(max_len + 4, 40)
                    )

                # Conditional color για status
                status_col_idx = None
                for idx, cell in enumerate(ws[1], start=1):
                    if cell.value == "status":
                        status_col_idx = idx
                        break

                if status_col_idx:
                    from openpyxl.styles import PatternFill as PF
                    green  = PF(start_color="D1FAE5", end_color="D1FAE5",
                                fill_type="solid")
                    red    = PF(start_color="FEE2E2", end_color="FEE2E2",
                                fill_type="solid")
                    yellow = PF(start_color="FEF9C3", end_color="FEF9C3",
                                fill_type="solid")
                    for row in ws.iter_rows(min_row=2):
                        cell = row[status_col_idx - 1]
                        if cell.value == "Validated":
                            cell.fill = green
                        elif cell.value == "Needs Human Review":
                            cell.fill = red
                        elif cell.value == "Pending":
                            cell.fill = yellow

            result.success      = True
            result.file_path    = out_path
            result.record_count = len(df)
            result.exported_at  = datetime.utcnow().isoformat()

        except Exception as e:
            result.error = str(e)

        return result

    # ── Εξαγωγή Line Items XLSX ─────────────────────────────────────────────

    def export_line_items_xlsx(self, records: List[Dict[str, Any]],
                               filename: str = None) -> ExportResult:
        """
        Εξαγωγή line items σε Excel. Κάθε line item γίνεται ξεχωριστή γραμμή,
        με τα βασικά πεδία του εγγράφου (αρχείο, προμηθευτής, ετικέτα κλπ)
        να επαναλαμβάνονται σε κάθε γραμμή.
        """
        result = ExportResult(format="xlsx")

        if not records:
            result.error = "Δεν υπάρχουν δεδομένα για εξαγωγή."
            return result

        try:
            rows = []
            for rec in records:
                # Βασικά πεδία εγγράφου
                base = {
                    "doc_id": rec.get("id", ""),
                    "filename": rec.get("filename", ""),
                    "label": rec.get("label", rec.get("schema_name", "")),
                    "supplier": rec.get("supplier", rec.get("vendor_name", "")),
                    "status": rec.get("status", ""),
                }
                # Scalar πεδία από result_json (π.χ. invoice_number, date, total)
                scalar_fields = {}
                for k, v in rec.items():
                    if k in ("id", "filename", "label", "schema_name", "supplier",
                             "vendor_name", "status", "result_json", "file_path",
                             "user_id", "confidence", "batch_id", "created_at",
                             "page_count", "page_index"):
                        continue
                    if isinstance(v, (str, int, float, type(None))):
                        scalar_fields[k] = v

                # Line items
                li = rec.get("line_items", [])
                if isinstance(li, str):
                    try:
                        import json as _json
                        li = _json.loads(li)
                    except:
                        li = []

                if li and isinstance(li, list):
                    for item in li:
                        if isinstance(item, dict):
                            row = {**base, **scalar_fields}
                            for ik, iv in item.items():
                                row[f"li_{ik}"] = iv
                            rows.append(row)
                else:
                    # Αν δεν υπάρχουν line items, βάλε μία γραμμή με τα scalars
                    rows.append({**base, **scalar_fields})

            if not rows:
                result.error = "Δεν βρέθηκαν line items."
                return result

            df = pd.DataFrame(rows)
            out_path = self._make_output_path(filename, "xlsx")

            with pd.ExcelWriter(str(out_path), engine="openpyxl") as writer:
                df.to_excel(writer, index=False, sheet_name="Line Items")

                from openpyxl.styles import Font, PatternFill, Alignment
                ws = writer.sheets["Line Items"]

                # Header styling
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(
                    start_color="059669", end_color="059669",
                    fill_type="solid"
                )
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")

                # Auto-width
                for col in ws.columns:
                    max_len = max(
                        (len(str(cell.value)) if cell.value else 0)
                        for cell in col
                    )
                    ws.column_dimensions[col[0].column_letter].width = (
                        min(max_len + 4, 40)
                    )

            result.success = True
            result.file_path = out_path
            result.record_count = len(df)
            result.exported_at = datetime.utcnow().isoformat()

        except Exception as e:
            result.error = str(e)

        return result

    # ── Εξαγωγή JSON ─────────────────────────────────────────────────────────

    def export_json(self, records: List[Dict[str, Any]],
                    filename: str = None) -> ExportResult:
        """Εξαγωγή records σε JSON αρχείο."""
        result = ExportResult(format="json")

        if not records:
            result.error = "Δεν υπάρχουν δεδομένα για εξαγωγή."
            return result

        try:
            out_path = self._make_output_path(filename, "json")
            out_path.write_text(
                json.dumps(records, ensure_ascii=False, indent=2,
                           default=str),
                encoding="utf-8"
            )
            result.success      = True
            result.file_path    = out_path
            result.record_count = len(records)
            result.exported_at  = datetime.utcnow().isoformat()

        except Exception as e:
            result.error = str(e)

        return result

    # ── Στατιστικά ───────────────────────────────────────────────────────────

    def summary_stats(self, records: List[Dict[str, Any]]) -> Dict[str, Any]:
        """
        Επιστρέφει στατιστικά για τα records:
        πλήθος ανά status, σύνολα ποσών, εύρος ημερομηνιών.
        """
        if not records:
            return {"total": 0}

        df = pd.DataFrame(records)

        stats: Dict[str, Any] = {"total": len(df)}

        # Πλήθος ανά status
        if "status" in df.columns:
            stats["by_status"] = df["status"].value_counts().to_dict()

        # Σύνολα ποσών
        for col in ["net_amount", "vat_amount", "total_amount"]:
            if col in df.columns:
                numeric = pd.to_numeric(df[col], errors="coerce")
                stats[f"sum_{col}"]  = round(float(numeric.sum()), 2)
                stats[f"mean_{col}"] = round(float(numeric.mean()), 2)

        # Εύρος ημερομηνιών
        if "invoice_date" in df.columns:
            dates = df["invoice_date"].dropna()
            if not dates.empty:
                stats["date_range"] = {
                    "from": str(dates.min()),
                    "to":   str(dates.max()),
                }

        return stats

    # ── Utilities ────────────────────────────────────────────────────────────

    def _to_dataframe(self, records: List[Dict],
                      columns: List[str] = None) -> pd.DataFrame:
        """Μετατροπή records σε DataFrame με ταξινόμηση στηλών."""
        df = pd.DataFrame(records)

        if columns:
            # Κρατάμε μόνο τις ζητούμενες στήλες (αν υπάρχουν)
            existing = [c for c in columns if c in df.columns]
            df = df[existing]
        else:
            # Ταξινόμηση: priority columns πρώτα, υπόλοιπες αλφαβητικά
            priority = [c for c in PRIORITY_COLUMNS if c in df.columns]
            rest     = sorted([c for c in df.columns if c not in priority])
            df = df[priority + rest]

        return df

    def _make_output_path(self, filename: str, ext: str) -> Path:
        """Δημιουργεί path για το αρχείο εξαγωγής."""
        if filename:
            # Διασφαλίζουμε σωστή κατάληξη
            p = Path(filename)
            if p.suffix.lower() != f".{ext}":
                p = p.with_suffix(f".{ext}")
            # Αν είναι απόλυτο path, το χρησιμοποιούμε ως έχει
            if p.is_absolute():
                return p
            return self.export_dir / p
        else:
            ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
            return self.export_dir / f"export_{ts}.{ext}"
